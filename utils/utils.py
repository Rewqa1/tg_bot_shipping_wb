import json
import os
from datetime import datetime

import aiofiles
from openpyxl.styles import Font
from openpyxl.workbook import Workbook


async def read_json(file_path, encoding='utf-8'):
    try:
        async with aiofiles.open(file_path, 'r', encoding=encoding) as file:
            content = await file.read()
            data = json.loads(content)
        return data
    except FileNotFoundError:
        print(f"File '{file_path}' not found ")
        return None
    except Exception as e:
        print(f"Error reading file '{file_path}': {e}")
        return None

async def create_incomes_excel(incomes_data, all_incomes_data):

    date_now = datetime.now().strftime("%d.%m.%Y %H:%M")

    directory = "excel_files"
    filename = "Поставки.xlsx"
    file_path = os.path.join(directory, filename)

    if not os.path.exists(file_path):
        os.makedirs(directory, exist_ok=False)


    wb = Workbook()
    wb.remove(wb.active)

    '''Create full stat sheet'''
    ws = wb.create_sheet(title='Полная статистика поставок')
    ws['A1'] = 'Артикул'
    ws['B1'] = 'Количество'
    row = 2

    for key, item in all_incomes_data.items():

        if key != 'fullQuantity':
            ws[f'A{row}'] = key
            ws[f'B{row}'] = item
            row += 1


    ws[f'A{row+1}'] = 'Полное количество'
    ws[f'B{row+1}'] = all_incomes_data['fullQuantity']
    ws[f'A{row+1}'].font = Font(bold=True)
    ws[f'B{row+1}'].font = Font(bold=True)
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15

    ''' Create stat by stock'''
    for warehouse in incomes_data:
        ws = wb.create_sheet(title=warehouse['warehouseName'])

        headers = ['Артикул', 'Количество']
        ws.append(headers)

        column_widths = {'A': 25, 'B': 15}
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        for article in warehouse['articles']:
            ws.append([article['article'] + article['size'], article['quantity']])

        max_row = len(warehouse['articles']) + 3
        ws[f'A{max_row}'] = 'Полное количество'
        ws[f'B{max_row}'] = warehouse['fullIncomeQuantityStock']
        ws[f'A{max_row}'].font = Font(bold=True)
        ws[f'B{max_row}'].font = Font(bold=True)

    wb.save(file_path)
